#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
数据导入模块（完整版）
- 读取 traits.xlsx
- 创建数据库表（包含序号、颜色）
- 提取图标并保存，每个特质一个图标，按序号命名
- 确保所有744条特质完整导入
"""

import os
import sqlite3
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Color

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
MOMO_DIR = os.path.join(BASE_DIR, "momodata")
DATA_DIR = os.path.join(MOMO_DIR, "data")
ICON_DIR = os.path.join(DATA_DIR, "icons")
DB_PATH = os.path.join(DATA_DIR, "leader_traits.db")

os.makedirs(DATA_DIR, exist_ok=True)
os.makedirs(ICON_DIR, exist_ok=True)

def ensure_trait_table():
    """确保特质表存在，如有必要添加颜色列"""
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS traits (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            number INTEGER UNIQUE,
            name TEXT NOT NULL,
            icon_path TEXT,
            ruler INTEGER DEFAULT 0,
            commander INTEGER DEFAULT 0,
            scientist INTEGER DEFAULT 0,
            profession_requirements TEXT,
            other_requirements TEXT,
            exclusive_with TEXT,
            attribute TEXT,
            council INTEGER DEFAULT 0,
            governor INTEGER DEFAULT 0,
            special TEXT,
            effect TEXT,
            effect_color TEXT,
            code TEXT UNIQUE NOT NULL
        )
    ''')
    # 检查是否已有 effect_color 列，若无则添加
    cursor.execute("PRAGMA table_info(traits)")
    columns = [col[1] for col in cursor.fetchall()]
    if 'effect_color' not in columns:
        cursor.execute("ALTER TABLE traits ADD COLUMN effect_color TEXT")
    conn.commit()
    conn.close()
    print("[导入] 特质表已确保存在")

def get_cell_color(cell):
    """获取单元格字体颜色，返回十六进制字符串或 None"""
    if cell.font and cell.font.color:
        color = cell.font.color
        if color.rgb:
            rgb = color.rgb
            # openpyxl 可能返回 RGB 对象或字符串
            if isinstance(rgb, str):
                if len(rgb) == 8:
                    return '#' + rgb[2:]  # 去掉透明度
                elif len(rgb) == 6:
                    return '#' + rgb
                else:
                    return None
            else:
                # 可能是 RGB 对象，尝试转换为十六进制
                try:
                    if hasattr(rgb, 'red') and hasattr(rgb, 'green') and hasattr(rgb, 'blue'):
                        return '#{:02x}{:02x}{:02x}'.format(rgb.red, rgb.green, rgb.blue)
                except:
                    pass
    return None

def extract_icons_and_colors_from_excel(excel_path):
    """提取图标和效果列颜色，返回 (行号到图标路径, 行号到颜色)"""
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb.active
    images = ws._images
    row_to_path = {}
    row_to_color = {}
    used_rows = set()

    # 先提取图标
    for idx, img in enumerate(images):
        anchor = img.anchor
        row = None
        if hasattr(anchor, '_from'):
            cell = anchor._from
            row = cell.row
        elif isinstance(anchor, str):
            import re
            match = re.search(r'\d+', anchor)
            if match:
                row = int(match.group())
        if row is None:
            print(f"[导入] 警告: 无法解析图片 {idx} 的行号，跳过")
            continue
        if row in used_rows:
            continue
        used_rows.add(row)
        temp_path = os.path.join(ICON_DIR, f"row_{row}.png")
        with open(temp_path, 'wb') as f:
            f.write(img._data())
        row_to_path[row] = temp_path

    # 提取效果列颜色（假设效果列是第15列，即列 O，索引 14）
    # 根据表头，效果列是第15列（O列），第一行是标题，数据从第2行开始
    max_row = ws.max_row
    for row in range(2, max_row + 1):
        cell = ws.cell(row, 15)  # O列
        color = get_cell_color(cell)
        if color:
            row_to_color[row] = color

    wb.close()
    return row_to_path, row_to_color

def import_from_excel(excel_path):
    if not os.path.exists(excel_path):
        print(f"[导入] Excel文件不存在: {excel_path}，跳过导入")
        return False

    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("DELETE FROM traits")
    conn.commit()
    print("[导入] 已清空旧数据")

    row_to_icon, row_to_color = extract_icons_and_colors_from_excel(excel_path)
    print(f"[导入] 提取到 {len(row_to_icon)} 个图标，{len(row_to_color)} 个颜色")

    try:
        df = pd.read_excel(excel_path, header=0)
        df.columns = [str(col).strip().replace('\n', '') for col in df.columns]
        print(f"[导入] 列名: {list(df.columns)}")

        required_cols = ['序号', '代码', '名称']
        for col in required_cols:
            if col not in df.columns:
                print(f"[导入] 错误: 缺少必要列 '{col}'")
                return False

        inserted = 0
        total_rows = len(df)

        for idx, row in df.iterrows():
            number = row.get('序号')
            if pd.isna(number):
                print(f"[导入] 跳过第 {idx+2} 行: 无序号")
                continue
            number = int(number)

            name = str(row.get('名称', '')).replace('\n', ' ').strip()
            if not name:
                print(f"[导入] 警告: 序号 {number} 名称为空")

            code = str(row.get('代码', '')).strip()
            if not code:
                print(f"[导入] 跳过序号 {number}: 无代码")
                continue

            excel_row = idx + 2
            icon_path = row_to_icon.get(excel_row, "")
            effect_color = row_to_color.get(excel_row, "")

            if icon_path and number:
                new_icon_path = os.path.join(ICON_DIR, f"{number}.png")
                if os.path.exists(icon_path):
                    os.rename(icon_path, new_icon_path)
                    icon_path = new_icon_path

            ruler = 1 if pd.notna(row.get('行政官')) and str(row.get('行政官')).strip() != '' else 0
            commander = 1 if pd.notna(row.get('指挥官')) and str(row.get('指挥官')).strip() != '' else 0
            scientist = 1 if pd.notna(row.get('科学家')) and str(row.get('科学家')).strip() != '' else 0
            council = 1 if pd.notna(row.get('内阁')) and str(row.get('内阁')).strip() != '' else 0
            governor = 1 if pd.notna(row.get('总督')) and str(row.get('总督')).strip() != '' else 0

            cursor.execute('''
                INSERT OR IGNORE INTO traits (
                    number, name, icon_path, ruler, commander, scientist,
                    profession_requirements, other_requirements, exclusive_with,
                    attribute, council, governor, special, effect, effect_color, code
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                number,
                name,
                icon_path,
                ruler,
                commander,
                scientist,
                str(row.get('职业需求', '')).strip(),
                str(row.get('其他需求', '')).strip(),
                str(row.get('互斥特质', '')).strip(),
                str(row.get('属性', '')).strip(),
                council,
                governor,
                str(row.get('特殊', '')).strip(),
                str(row.get('效果', '')).strip(),
                effect_color,
                code
            ))
            inserted += 1

        conn.commit()
        print(f"[导入] 成功插入 {inserted} 条特质")
        cursor.execute("SELECT COUNT(*) FROM traits")
        total = cursor.fetchone()[0]
        print(f"[导入] 数据库当前总记录数: {total}")
        return True
    except Exception as e:
        print(f"[导入] 失败: {e}")
        return False
    finally:
        conn.close()

def is_database_empty():
    if not os.path.exists(DB_PATH):
        return True
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("SELECT COUNT(*) FROM traits")
    count = cursor.fetchone()[0]
    conn.close()
    return count == 0

def need_icons():
    if not os.path.exists(ICON_DIR):
        return True
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("SELECT COUNT(*) FROM traits WHERE icon_path != '' AND icon_path IS NOT NULL")
    count_with_icon = cursor.fetchone()[0]
    conn.close()
    existing_icons = len([f for f in os.listdir(ICON_DIR) if f.endswith('.png')])
    return existing_icons < count_with_icon

def run_import(excel_path):
    ensure_trait_table()
    if is_database_empty() or need_icons():
        print("[导入] 检测到数据库为空或缺少图标，开始导入...")
        return import_from_excel(excel_path)
    else:
        print("[导入] 数据库已有数据且图标存在，跳过导入")
        return False

if __name__ == "__main__":
    excel_path = os.path.join(BASE_DIR, "traits.xlsx")
    run_import(excel_path)
